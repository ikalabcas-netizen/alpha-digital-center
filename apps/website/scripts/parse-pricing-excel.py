#!/usr/bin/env python3
"""
Parse contex/BẢNG GIÁ ALPHA DIGITAL.xlsx → prisma/seed-products.sql

Excel structure:
- Row 8: header row (MÃ SP, DÒNG SP/DV, VẬT LIỆU, Chi tiết, ĐVT, GIÁ, GHI CHÚ)
- Section headers: col A has text, cols B–F empty (e.g. "DỊCH VỤ GIA CÔNG")
  → become ProductCategory
- Parent rows: col B (DÒNG SP/DV) non-null → new Product in current category.
  If parent row also has col D+F populated, it contains the first variant inline.
- Child rows: col A (MÃ SP) + col D (Chi tiết) + col F (GIÁ) non-null → ProductVariant
  of the current parent.

Emits SQL with:
- INSERT ProductCategory (ON CONFLICT slug DO NOTHING)
- INSERT Product (ON CONFLICT slug DO NOTHING)
- INSERT ProductVariant (no ON CONFLICT — fresh insert with new uuid)
- INSERT BlogPost (ON CONFLICT slug DO NOTHING) — 1 post per variant, category='san-pham'

Usage:
  python scripts/parse-pricing-excel.py [--limit N] [--out path]
    --limit N   : emit only first N Products (and their variants/blog posts)
    --out PATH  : output file (default: prisma/seed-products.sql)
"""
import argparse
import io
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
EXCEL = ROOT / 'contex' / 'BẢNG GIÁ ALPHA DIGITAL.xlsx'


def slugify(s: str) -> str:
    s = (s or '').lower().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd')
    s = re.sub(r'[^a-z0-9]+', '-', s)
    s = re.sub(r'^-+|-+$', '', s)
    return s[:90]


def sql_str(v) -> str:
    """SQL literal for TEXT: NULL if None/empty, else 'escaped' string."""
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s:
        return 'NULL'
    return "'" + s.replace("'", "''") + "'"


def sql_num(v) -> str:
    """SQL literal for NUMERIC/BIGINT — accepts int, float, str with commas/dots."""
    if v is None:
        return 'NULL'
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).strip().replace('.', '').replace(',', '').replace(' ', '')
    if not s:
        return 'NULL'
    try:
        return str(int(float(s)))
    except Exception:
        return 'NULL'


def format_price(n: int) -> str:
    return f"{n:,}".replace(',', '.')


def description_for_variant(product_name: str, variant_name: str, material: str, unit: str, price: int, note: str) -> dict:
    """Generate BlogPost fields (excerpt, content in Markdown, SEO) for a variant."""
    price_fmt = format_price(price)
    note_clean = (note or '').strip()

    title = f"{product_name} — {variant_name}"
    excerpt_parts = [f"{variant_name} thuộc dòng {product_name} của Alpha Digital Center."]
    excerpt_parts.append(f"Giá tham khảo: {price_fmt} VND/{unit}.")
    excerpt = ' '.join(excerpt_parts)

    mat_line = f"**Vật liệu:** {material}" if material else ""
    note_para = f"> {note_clean}" if note_clean else ""

    content_lines = [
        f"## {variant_name} — chi tiết dịch vụ",
        "",
        f"**{variant_name}** là một biến thể thuộc dòng sản phẩm/dịch vụ **{product_name}** do Alpha Digital Center gia công. Biến thể này phù hợp với các yêu cầu gia công đặc thù trong phòng lab nha khoa hiện đại.",
        "",
        "### Thông số kỹ thuật",
        "",
    ]
    if material:
        content_lines.append(f"- **Vật liệu:** {material}")
    content_lines.extend([
        f"- **Đơn vị tính:** {unit}",
        f"- **Giá tham khảo:** {price_fmt} VND / {unit}",
    ])
    if note_clean:
        content_lines.extend(["", "### Ghi chú từ nhà sản xuất", "", f"> {note_clean}"])
    content_lines.extend([
        "",
        "### Quy trình gia công tại Alpha Digital Center",
        "",
        "Mỗi biến thể gia công tại ADC đều trải qua quy trình khép kín, đảm bảo chất lượng ổn định và thời gian giao đúng hẹn:",
        "",
        "1. **Tiếp nhận file & kiểm tra** — đội QC kiểm tra file CAD/CAM do labo gửi về, báo cáo sai sót nếu có.",
        "2. **Thiết kế/điều chỉnh** (nếu thuê) — các chuyên viên thiết kế của ADC hiệu chỉnh theo vật liệu và yêu cầu lâm sàng.",
        "3. **Gia công chính xác** — hệ thống máy CNC 5 trục, in 3D kim loại SLM, scanner E4 đảm bảo dung sai tối thiểu.",
        "4. **QC xuất xưởng** — 5 vòng kiểm tra trước khi đóng gói, kèm phiếu gia công ghi rõ lô vật liệu.",
        "",
        "### Ứng dụng lâm sàng",
        "",
        f"{variant_name} thường được chỉ định cho các case phục hình có yêu cầu độ chính xác cao và tính thẩm mỹ tự nhiên. Labo đối tác của Alpha Digital Center có thể đặt số lượng lớn với lead time ổn định.",
        "",
        "### Liên hệ đặt hàng",
        "",
        f"Xem thêm các biến thể khác trong dòng [{product_name}](/san-pham/{{PRODUCT_SLUG}}) hoặc [liên hệ trực tiếp](/lien-he) để nhận báo giá chi tiết cho số lượng lớn và lịch giao hàng.",
        "",
    ])
    content = '\n'.join(content_lines)

    seo_title = f"{title} | Giá {price_fmt} VND — Alpha Digital Center"[:160]
    seo_desc = (excerpt + ' Bảo hành minh bạch, giao hàng đúng hẹn từ ADC.')[:160]

    tags = []
    if material:
        tags.append(material)
    tags.append(variant_name)

    return {
        'title': title,
        'excerpt': excerpt,
        'content': content,
        'seo_title': seo_title,
        'seo_description': seo_desc,
        'tags': tags,
    }


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--limit', type=int, default=0, help='limit number of Products (0 = all)')
    p.add_argument('--out', default='prisma/seed-products.sql')
    args = p.parse_args()

    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb.active

    categories = {}  # slug -> {nameVi, display_order}
    products = []    # list of dicts
    current_cat_slug = None
    current_product_idx = None

    def ensure_category(name_vi: str) -> str:
        slug = slugify(name_vi)
        if slug not in categories:
            categories[slug] = {'nameVi': name_vi, 'displayOrder': len(categories) + 1}
        return slug

    def new_product(name_vi: str, material: str, first_variant_row, cat_slug: str) -> int:
        slug_base = slugify(name_vi)
        slug = slug_base
        i = 2
        while any(p['slug'] == slug for p in products):
            slug = f"{slug_base}-{i}"
            i += 1

        sku_first = first_variant_row.get('sku') if first_variant_row else None
        variants = []
        if first_variant_row and first_variant_row.get('nameVi') and first_variant_row.get('price') is not None:
            variants.append(first_variant_row)

        products.append({
            'nameVi': name_vi,
            'slug': slug,
            'sku': sku_first,
            'material': material,
            'category_slug': cat_slug,
            'displayOrder': len(products) + 1,
            'variants': variants,
        })
        return len(products) - 1

    def add_variant(product_idx: int, variant: dict):
        products[product_idx]['variants'].append(variant)

    row = 9
    while row <= ws.max_row:
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        c = ws.cell(row, 3).value
        d = ws.cell(row, 4).value
        e = ws.cell(row, 5).value  # ĐVT
        f = ws.cell(row, 6).value  # GIÁ
        g = ws.cell(row, 7).value  # GHI CHÚ
        # Section header: A has text, B-F all empty
        if a and all(ws.cell(row, col).value is None for col in range(2, 7)):
            current_cat_slug = ensure_category(str(a).strip())
            current_product_idx = None
            row += 1
            continue
        # Parent row (new Product)
        if b:
            if current_cat_slug is None:
                current_cat_slug = ensure_category('Chưa phân loại')
            first_variant = None
            if d and f is not None:
                first_variant = {
                    'sku': str(a).strip() if a else None,
                    'nameVi': str(d).strip(),
                    'material': str(c).strip() if c else None,
                    'unit': (str(e).strip() if e else 'Cái'),
                    'price': int(float(sql_num(f))) if sql_num(f) != 'NULL' else None,
                    'note': (str(g).strip() if g else None),
                }
            current_product_idx = new_product(
                name_vi=str(b).strip(),
                material=(str(c).strip() if c else None),
                first_variant_row=first_variant,
                cat_slug=current_cat_slug,
            )
            row += 1
            continue
        # Child row (variant of current parent)
        if a and d and f is not None and current_product_idx is not None:
            variant = {
                'sku': str(a).strip(),
                'nameVi': str(d).strip(),
                'material': str(c).strip() if c else None,
                'unit': (str(e).strip() if e else 'Cái'),
                'price': int(float(sql_num(f))) if sql_num(f) != 'NULL' else None,
                'note': (str(g).strip() if g else None),
            }
            if variant['price'] is not None:
                add_variant(current_product_idx, variant)
            row += 1
            continue
        # Orphan or blank — skip
        row += 1

    # Apply limit
    if args.limit and args.limit > 0:
        products = products[:args.limit]
        # keep only categories referenced by limited products
        used_cats = {p['category_slug'] for p in products}
        categories = {k: v for k, v in categories.items() if k in used_cats}

    total_variants = sum(len(p['variants']) for p in products)
    total_blog = total_variants

    # ====== Emit SQL ======
    out_path = ROOT / args.out
    out_path.parent.mkdir(parents=True, exist_ok=True)

    buf = io.StringIO()
    w = buf.write

    w('-- Auto-generated by scripts/parse-pricing-excel.py\n')
    w(f'-- Source: contex/BẢNG GIÁ ALPHA DIGITAL.xlsx\n')
    w(f'-- Stats: {len(categories)} categories, {len(products)} products, {total_variants} variants, {total_blog} blog posts\n')
    w(f"-- Run: docker exec -i <db-container> psql -U adc_admin -d adc_marketing < {args.out}\n")
    w('\nBEGIN;\n\n')

    # Categories
    w('-- ========== Categories ==========\n')
    for slug, cat in categories.items():
        w(
            'INSERT INTO product_categories (id, name_vi, slug, display_order, is_active, created_at, updated_at) VALUES '
            f"(gen_random_uuid(), {sql_str(cat['nameVi'])}, {sql_str(slug)}, {cat['displayOrder']}, true, NOW(), NOW()) "
            'ON CONFLICT (slug) DO NOTHING;\n'
        )
    w('\n')

    # Products
    w('-- ========== Products ==========\n')
    for prod in products:
        variants = prod['variants']
        desc_lines = []
        if variants:
            desc_lines.append(f"Dịch vụ bao gồm {len(variants)} biến thể:")
            for v in variants[:6]:
                desc_lines.append(f"- {v['nameVi']}")
            if len(variants) > 6:
                desc_lines.append(f"- ... và {len(variants) - 6} biến thể khác")
            desc_lines.append('')
            desc_lines.append('Liên hệ Alpha Digital Center để nhận báo giá chi tiết theo số lượng.')
        description = '\n'.join(desc_lines) if desc_lines else None

        seo_title = f"{prod['nameVi']} — Alpha Digital Center"
        seo_desc = (description[:160].replace('\n', ' ') if description else f"{prod['nameVi']} tại Alpha Digital Center.")

        # Subquery to get category id by slug
        cat_id_expr = f"(SELECT id FROM product_categories WHERE slug = {sql_str(prod['category_slug'])})"

        w(
            'INSERT INTO products '
            '(id, category_id, sku, name_vi, slug, description_vi, material, display_order, is_active, is_featured, seo_title, seo_description, created_at, updated_at) VALUES '
            f"(gen_random_uuid(), {cat_id_expr}, {sql_str(prod['sku'])}, {sql_str(prod['nameVi'])}, {sql_str(prod['slug'])}, "
            f"{sql_str(description)}, {sql_str(prod['material'])}, {prod['displayOrder']}, true, false, "
            f"{sql_str(seo_title)}, {sql_str(seo_desc)}, NOW(), NOW()) "
            'ON CONFLICT (slug) DO NOTHING;\n'
        )
    w('\n')

    # Variants + BlogPosts
    w('-- ========== ProductVariants + BlogPosts (one post per variant) ==========\n')
    for prod in products:
        prod_id_expr = f"(SELECT id FROM products WHERE slug = {sql_str(prod['slug'])})"
        for order, v in enumerate(prod['variants'], start=1):
            # Variant
            w(
                'INSERT INTO product_variants '
                '(id, product_id, variant_name_vi, unit, price_vnd, price_note, is_active, created_at) VALUES '
                f"(gen_random_uuid(), {prod_id_expr}, {sql_str(v['nameVi'])}, {sql_str(v['unit'])}, "
                f"{v['price']}, {sql_str(v.get('note'))}, true, NOW());\n"
            )

            # BlogPost for SEO
            blog = description_for_variant(
                product_name=prod['nameVi'],
                variant_name=v['nameVi'],
                material=v.get('material') or prod['material'] or '',
                unit=v['unit'],
                price=v['price'],
                note=v.get('note') or '',
            )
            blog_slug_base = slugify(f"{prod['slug']}-{v['nameVi']}")
            # replace PRODUCT_SLUG placeholder in content
            blog_content = blog['content'].replace('{PRODUCT_SLUG}', prod['slug'])

            # Tags as Postgres array literal
            tags_array = 'ARRAY[' + ','.join(sql_str(t) for t in blog['tags']) + ']::text[]' if blog['tags'] else "'{}'::text[]"

            w(
                'INSERT INTO blog_posts '
                '(id, title_vi, slug, excerpt_vi, content_vi, category, tags, status, published_at, '
                'seo_title, seo_description, view_count, ai_generated, created_at, updated_at) VALUES '
                f"(gen_random_uuid(), {sql_str(blog['title'])}, {sql_str(blog_slug_base)}, "
                f"{sql_str(blog['excerpt'])}, {sql_str(blog_content)}, 'san-pham', "
                f"{tags_array}, 'published', NOW(), "
                f"{sql_str(blog['seo_title'])}, {sql_str(blog['seo_description'])}, 0, false, NOW(), NOW()) "
                'ON CONFLICT (slug) DO NOTHING;\n'
            )
        w('\n')

    w('COMMIT;\n\n')
    w('-- Summary counts\n')
    w(
        "SELECT 'categories' AS t, count(*) FROM product_categories "
        "UNION ALL SELECT 'products', count(*) FROM products "
        "UNION ALL SELECT 'variants', count(*) FROM product_variants "
        "UNION ALL SELECT 'blog_posts_san_pham', count(*) FROM blog_posts WHERE category = 'san-pham';\n"
    )

    out_path.write_text(buf.getvalue(), encoding='utf-8')

    sys.stdout.reconfigure(encoding='utf-8')
    print(f"[ok] wrote {out_path}")
    print(f"     {len(categories)} categories, {len(products)} products, {total_variants} variants, {total_blog} blog posts")


if __name__ == '__main__':
    main()
